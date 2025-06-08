from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from datetime import datetime
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
import os
import time
import tkinter as tk
from tkinter import messagebox

def coletar_dados():
    # Configura o Chrome em modo invisível (headless)
    options = Options()
    options.add_argument('--headless')
    navegador = webdriver.Chrome(options=options)
    navegador.get("https://www.climatempo.com.br/previsao-do-tempo/cidade/558/saopaulo-sp")

    time.sleep(5)  # Aguarda o carregamento da página
    html = navegador.page_source
    navegador.quit()

    # Processa o HTML com BeautifulSoup
    soup = BeautifulSoup(html, 'html.parser')

    # Captura dados das metatags
    tmin = soup.find("meta", {"name": "tmin"})
    tmax = soup.find("meta", {"name": "tmax"})
    urmax = soup.find("meta", {"name": "urmax"})

    temperatura = f"{tmin['content']}ºC / {tmax['content']}ºC" if tmin and tmax else "Não encontrada"
    umidade = f"{urmax['content']}%" if urmax else "Não encontrada"
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Grava os dados no Excel
    arquivo = "dados_climaticos.xlsx"
    if os.path.exists(arquivo):
        wb = load_workbook(arquivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Data/Hora", "Temperatura", "Umidade"])

    ws.append([agora, temperatura, umidade])
    wb.save(arquivo)

    return f"Registrado: {agora}\nTemperatura: {temperatura}\nUmidade: {umidade}"

# Interface gráfica com Tkinter
def executar():
    msg = coletar_dados()
    messagebox.showinfo("Sucesso", msg)

janela = tk.Tk()
janela.title("Captador de Temperatura - SP")
janela.geometry("300x150")

label = tk.Label(janela, text="Clique no botão para captar os dados")
label.pack(pady=10)

botao = tk.Button(janela, text="Buscar previsão", command=executar)
botao.pack(pady=10)

janela.mainloop()
